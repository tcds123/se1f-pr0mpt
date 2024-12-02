# -*- coding: utf-8 -*-
import base64
import hashlib
import hmac
import json
import random
import string
import sys
import time
import uuid

import openpyxl
from openpyxl import Workbook

import requests

CHAR_LIST = []
[[CHAR_LIST.append(e) for e in string.ascii_letters] for i in range(0, 2)]
[[CHAR_LIST.append(e) for e in string.ascii_letters] for i in range(0, 2)]
[[CHAR_LIST.append(e) for e in string.digits] for i in range(0, 2)]


# generate nonce
def get_chars(length):
    """get the random chars

    :param length: lenght of the string
    :returns: string

    """
    random.shuffle(CHAR_LIST)
    return ''.join(CHAR_LIST[0:length])


def queryByOpenAPI(queryDict, secretKey, publicKey, remoteUrl):
    nonce = get_chars(8)
    timestamp = str(int(time.time() * 1000))
    unionStr = nonce + timestamp + json.dumps(queryDict, ensure_ascii=False)
    # print("unionStr:", unionStr)
    signature = base64.b64encode(
        hmac.new(secretKey.encode('utf-8'), unionStr.encode('utf-8'), digestmod=hashlib.sha256).digest()).decode('utf-8')
    # print("Signature: ", signature)
    headers = {
        "Content-Type": "application/json",
        "PublicKey": publicKey,
        "Nonce": nonce,
        "Timestamp": timestamp,
        "Signature": signature
    }
    data = requests.request("POST", remoteUrl, headers=headers,
                            data=str(json.dumps(queryDict, ensure_ascii=False)).encode('utf-8')).json()
    # print("response data:", data)
    return data

def print_2_xlsx():
    wb = openpyxl.load_workbook('D:/Users/zongwx1/Desktop/test_CC空调服务政策.xlsx')
    ws = wb.active
    for row_index, i in enumerate(ws.iter_rows(min_row=2, max_row=30, min_col=1, max_col=2), start=2):
        query = i[1].value + '\n' + i[0].value
        response = get_api_response(query)
        cell = ws.cell(row_index, 3)
        cell.value = response
        print(response)
        print("======")

# API URLs
final_url = "https://aibuilder.midea.com/aigc-openapi/v1/open/qa"

def get_api_response(query):
    # get uuid
    my_uuid = str(uuid.uuid1()).replace('-', '')

    ## 如下两个ID在"应用设置"中看到，替换为自己应用的ID
    pubkey = "27d829617f0848b3acc3d1fe2dd485fb"
    secret = "b89c5560cd2f43c59f21635450a1b7ad"

    # query = "你好"
    payload = {
        ## 可以替换为自己的IMIP名字
        "userId": "zongwx1",
        "requestId": my_uuid,
        "query": query,
        "customParam":
            {
                "temperature": 0.0,
            }
    }
    result = queryByOpenAPI(payload, secret, pubkey, final_url)

    return result['result']['llmOutput']

if __name__ == "__main__":
    print_2_xlsx()

    sys.exit(0)